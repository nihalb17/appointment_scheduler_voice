"""Document parser for RAG.

Extracts text from various document formats for knowledge base processing.
"""

import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


def parse_document(file_path: Path) -> str:
    """Parse a document and extract text content.
    
    Supports: PDF, DOCX, XLSX, CSV, TXT
    
    Args:
        file_path: Path to the document file.
        
    Returns:
        Extracted text content.
    """
    extension = file_path.suffix.lower()
    
    try:
        if extension == ".pdf":
            return _parse_pdf(file_path)
        elif extension == ".docx":
            return _parse_docx(file_path)
        elif extension == ".xlsx":
            return _parse_xlsx(file_path)
        elif extension == ".csv":
            return _parse_csv(file_path)
        elif extension == ".txt":
            return _parse_txt(file_path)
        else:
            logger.warning(f"Unsupported file format: {extension}")
            return f"[Unsupported format: {extension}]"
    except Exception as e:
        logger.error(f"Error parsing {file_path}: {e}")
        return f"[Error parsing file: {e}]"


def _parse_pdf(file_path: Path) -> str:
    """Extract text from PDF file."""
    try:
        import PyPDF2
        text = []
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text.append(page.extract_text())
        return "\n".join(text)
    except ImportError:
        logger.warning("PyPDF2 not installed, cannot parse PDF")
        return "[PDF parsing requires PyPDF2]"


def _parse_docx(file_path: Path) -> str:
    """Extract text from DOCX file."""
    try:
        from docx import Document
        doc = Document(file_path)
        text = []
        for para in doc.paragraphs:
            text.append(para.text)
        return "\n".join(text)
    except ImportError:
        logger.warning("python-docx not installed, cannot parse DOCX")
        return "[DOCX parsing requires python-docx]"


def _parse_xlsx(file_path: Path) -> str:
    """Extract text from XLSX file."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        text = []
        for sheet in wb.worksheets:
            text.append(f"--- Sheet: {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    text.append(row_text)
        return "\n".join(text)
    except ImportError:
        logger.warning("openpyxl not installed, cannot parse XLSX")
        return "[XLSX parsing requires openpyxl]"


def _parse_csv(file_path: Path) -> str:
    """Extract text from CSV file."""
    import csv
    text = []
    with open(file_path, 'r', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            text.append(" | ".join(row))
    return "\n".join(text)


def _parse_txt(file_path: Path) -> str:
    """Extract text from TXT file."""
    return file_path.read_text(encoding='utf-8')
